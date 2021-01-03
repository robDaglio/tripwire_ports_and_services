#!python
import os
from openpyxl import load_workbook

class PasData:
    def __init__(self, path, min_rc):
        self.path = path
        self.min_rc = min_rc
        self.files = {x: os.path.join(os.path.abspath(self.path), x) for x in os.listdir(self.path)}
        self.data = dict()
        
    def read_sheets(self):
        for file, path in self.files.items():
            work_book = load_workbook(filename=path)
            active = work_book.active
            total_rows, total_cols = len(active['A']), len(active['1'])

            wb_data = list()

            for row in active.iter_rows(
                min_row=int(self.min_rc[0]),
                min_col=int(self.min_rc[1]),
                max_row=total_rows,
                max_col=total_cols,
            ):
                d = list()
                for cell in row:
                    d.append(cell.value)
                wb_data.append(d)
            
            self.data[file] = wb_data

def match(ci_function_data, pas_data):
    ci_mapped = dict()
    for file_name, data_list in pas_data.items():
        if file_name in ci_function_data.keys():
            for ci in ci_function_data[file_name]:
                ci_mapped[ci] = pas_data[file_name]
    return ci_mapped
    

        